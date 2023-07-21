import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_or_append_excel():
    filename = "Client List (Please add month here) 2023.xlsx"
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        row_count = worksheet.max_row + 1
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['#', 'LHW Code', "LHW Name", 'WI Woman Code', 'Khandan #', 'Client Name', 'Received Method Name', 'Method Start Date'])
        row_count = 2
    
    lhw_code = lhw_code_entry.get()
    lhw_name = lhw_name_entry.get()
    wi_woman_code = wi_woman_code_entry.get()
    khandan_number = khandan_number_entry.get()
    client_name = client_name_entry.get()
    dropdown_value = dropdown.get()
    start_date = start_date_entry.get()

    if not lhw_code.isdigit() or not wi_woman_code.isdigit() or not khandan_number.isdigit():
        messagebox.showerror("Invalid Input", "LHW Code, WI Woman Code, and Khandan # should be numbers.")
        return

    if not lhw_name.replace(" ", "").isalpha():
        messagebox.showerror("Invalid Input", "LHW Name should only contain alphabetic characters and spaces.")
        return

    if not client_name.replace(" ", "").isalpha():
        messagebox.showerror("Invalid Input", "Client Name should only contain alphabetic characters and spaces.")
        return

    if not all([lhw_code, lhw_name, wi_woman_code, khandan_number, client_name, dropdown_value, start_date]):
        messagebox.showerror("Missing Input", "Please fill in all the required fields.")
        return
    dropdown_value = dropdown.get()
    
    if dropdown_value == "Please Select a Method":
        messagebox.showerror("Invalid Input", "Please select a valid method.")
        return

    try:
        datetime.strptime(start_date, "%d/%m/%Y")
    except ValueError:
        messagebox.showerror("Invalid Input", "Method Start Date should be in the format DD/MM/YYYY.")
        return

    worksheet.append([row_count - 1, lhw_code, lhw_name, wi_woman_code, khandan_number, client_name, dropdown_value, start_date])
    
    last_row = worksheet.max_row
    last_column = worksheet.max_column

    border_thickness = 1.5

    for row in worksheet.iter_rows(min_row=1, max_row=last_row, max_col=last_column):
        for cell in row:
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

    header_fill = PatternFill(start_color="B7CCE1", end_color="B7CCE1", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = header_fill

    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(filename)
    messagebox.showinfo("Success", "Data has been saved to the Excel file.")
    
    lhw_code_entry.delete(0, tk.END)
    lhw_name_entry.delete(0, tk.END)
    wi_woman_code_entry.delete(0, tk.END)
    khandan_number_entry.delete(0, tk.END)
    client_name_entry.delete(0, tk.END)
    start_date_entry.delete(0, tk.END)
    
def create_ui():
    global lhw_code_entry, lhw_name_entry, wi_woman_code_entry, khandan_number_entry, client_name_entry, dropdown, start_date_entry
    
    root = tk.Tk()
    root.title("Willows Intl. PK - Client List Data Entry")
    root.geometry("450x350")
    
    main_frame = tk.Frame(root)
    main_frame.pack(padx=30, pady=30)
    
    lhw_code_label = tk.Label(main_frame, text="LHW Code:")
    lhw_code_label.grid(row=0, column=0, sticky="e")
    lhw_code_entry = tk.Entry(main_frame, width=30)
    lhw_code_entry.grid(row=0, column=1, padx=10, pady=5)
    
    lhw_name_label = tk.Label(main_frame, text="LHW Name:")
    lhw_name_label.grid(row=1, column=0, sticky="e")
    lhw_name_entry = tk.Entry(main_frame, width=30)
    lhw_name_entry.grid(row=1, column=1, padx=10, pady=5)
    
    wi_woman_code_label = tk.Label(main_frame, text="WI Woman Code:")
    wi_woman_code_label.grid(row=2, column=0, sticky="e")
    wi_woman_code_entry = tk.Entry(main_frame, width=30)
    wi_woman_code_entry.grid(row=2, column=1, padx=10, pady=5)
    
    khandan_number_label = tk.Label(main_frame, text="Khandan #:")
    khandan_number_label.grid(row=3, column=0, sticky="e")
    khandan_number_entry = tk.Entry(main_frame, width=30)
    khandan_number_entry.grid(row=3, column=1, padx=10, pady=5)
    
    client_name_label = tk.Label(main_frame, text="Client Name:")
    client_name_label.grid(row=4, column=0, sticky="e")
    client_name_entry = tk.Entry(main_frame, width=30)
    client_name_entry.grid(row=4, column=1, padx=10, pady=5)
    
    dropdown_label = tk.Label(main_frame, text="Received Method Name")
    dropdown_label.grid(row=5, column=0, sticky="e")
    dropdown = tk.StringVar(main_frame)
    dropdown.set("Please Select a Method")
    dropdown_menu = tk.OptionMenu(main_frame, dropdown, "Condoms","Pills", "Injection", "Implant", "IUCD", "PPIUCD", "TL")
    dropdown_menu.config(width=28)
    dropdown_menu.grid(row=5, column=1, padx=10, pady=5)
    
    start_date_label = tk.Label(main_frame, text="Method Start Date:")
    start_date_label.grid(row=6, column=0, sticky="e")
    start_date_entry = tk.Entry(main_frame, width=30)
    start_date_entry.grid(row=6, column=1, padx=10, pady=5)
    
    submit_button = tk.Button(main_frame, text="Submit", command=create_or_append_excel)
    submit_button.grid(row=7, column=0, columnspan=2, pady=10)
    
    developed_by_label = tk.Label(root, text="Developed by Mutahar (Willows International, Pakistan)", fg="gray", font=("Arial", 8))
    developed_by_label.pack(side="bottom")
    
    root.mainloop()

create_ui()
